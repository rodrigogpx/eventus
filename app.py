from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from flask_sqlalchemy import SQLAlchemy
from database.models import db, Participant, Event, Meeting, MeetingForm, participant_event, Attendance  # Import Attendance
from flask_wtf import FlaskForm, CSRFProtect
from functools import wraps
import csv
from io import StringIO
from datetime import date, datetime, timedelta, time
from wtforms import StringField, EmailField, TextAreaField, IntegerField, validators, SelectMultipleField, PasswordField, DateField, TimeField, FieldList, FormField, SelectField
import os
from dotenv import load_dotenv
from config.menu import get_admin_menu  # Importar a configuração do menu
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from flask import send_file

# Carrega as variáveis de ambiente do arquivo .env
load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your_secret_key')

# Configuração do banco de dados
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
csrf = CSRFProtect(app)

@app.before_first_request
def create_tables():
    db.create_all()

# Decorator to require login for admin routes
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            flash('Por favor, faça login para acessar esta página.', 'error')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

class RegistrationForm(FlaskForm):
    full_name = StringField('Nome Completo', [validators.InputRequired(), validators.Length(max=100)])
    email = EmailField('Email', [validators.InputRequired(), validators.Email(), validators.Length(max=100)])
    phone = StringField('Telefone', [validators.InputRequired(), validators.Length(max=20)])
    city = StringField('Cidade', [validators.InputRequired(), validators.Length(max=50)])

class CheckinForm(FlaskForm):
    email = EmailField('Email', [validators.InputRequired(), validators.Email(), validators.Length(max=100)])

class MeetingForm(FlaskForm):
    title = StringField('Título', [validators.DataRequired()])
    date = DateField('Data', [validators.DataRequired()])
    time = TimeField('Hora', [validators.DataRequired()])
    description = TextAreaField('Descrição', [validators.Optional()])
    event_id = SelectField('Evento', coerce=int)

class EventForm(FlaskForm):
    theme = StringField('Tema', [validators.DataRequired()])
    date = DateField('Data', [validators.DataRequired()])
    max_participants = IntegerField('Máximo de Participantes', [validators.Optional()])
    description = TextAreaField('Descrição', [validators.Optional()])
    meetings = FieldList(FormField(MeetingForm), min_entries=1)

class LoginForm(FlaskForm):
    username = StringField('Nome de Usuário', [validators.DataRequired()])
    password = PasswordField('Senha', [validators.DataRequired()])

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    form = RegistrationForm()
    if form.validate_on_submit():
        if Participant.query.filter_by(email=form.email.data).first():
            flash('Email já registrado! Tente usar um email diferente ou entre em contato conosco para assistência.', 'error')
            return redirect(url_for('register'))

        new_participant = Participant(
            full_name=form.full_name.data,
            email=form.email.data,
            phone=form.phone.data,
            city=form.city.data
        )
        db.session.add(new_participant)
        db.session.commit()

        flash('Inscrição realizada com sucesso! Você receberá um email de confirmação em breve.', 'success')
        return redirect(url_for('confirmation'))

    return render_template('register.html', form=form)

@app.route('/confirmation')
def confirmation():
    return render_template('confirmation.html')

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    form = LoginForm()
    if form.validate_on_submit():
        if form.username.data == os.getenv('ADMIN_USERNAME') and form.password.data == os.getenv('ADMIN_PASSWORD'):
            session['logged_in'] = True
            session['username'] = form.username.data  # Armazena o nome do usuário na sessão
            flash('Login realizado com sucesso!', 'success')
            return redirect(url_for('admin_dashboard'))
        flash('Usuário ou senha inválidos!', 'error')
    return render_template('admin_login.html', form=form)

@app.route('/admin/logout')
def admin_logout():
    session.pop('logged_in', None)
    session.pop('username', None)  # Remove o nome do usuário da sessão
    flash('Logout realizado com sucesso!', 'success')
    return redirect(url_for('admin_login'))

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    # Estatísticas gerais
    stats = {
        'total_events': Event.query.count(),
        'total_participants': Participant.query.count(),
        'total_meetings': Meeting.query.count(),
        'upcoming_events': Event.query.filter(Event.date >= datetime.now()).count(),
        'completed_events': Event.query.filter(Event.date < datetime.now()).count(),
        'upcoming_meetings': Meeting.query.filter(Meeting.date >= datetime.now()).count(),
        'completed_meetings': Meeting.query.filter(Meeting.date < datetime.now()).count(),
        'new_participants': Participant.query.filter(
            Participant.created_at >= datetime.now() - timedelta(days=7)
        ).count(),
    }

    # Calcular média de participantes por evento
    if stats['total_events'] > 0:
        stats['avg_participants_per_event'] = round(stats['total_participants'] / stats['total_events'], 1)
    else:
        stats['avg_participants_per_event'] = 0

    # Buscar eventos recentes e futuros
    recent_events = Event.query.filter(Event.date >= datetime.now()).order_by(Event.date.asc()).limit(5).all()

    # Dados para os gráficos de pizza por cidade
    events_data = []
    events = Event.query.filter(Event.date >= datetime.now()).order_by(Event.date.asc()).all()
    
    for event in events:
        # Buscar todos os participantes do evento
        participants = db.session.query(Participant)\
            .join(participant_event)\
            .filter(participant_event.c.event_id == event.id)\
            .all()
        
        # Agrupar participantes por cidade
        cities_count = {}
        for participant in participants:
            cities_count[participant.city] = cities_count.get(participant.city, 0) + 1
        
        # Preparar dados para o gráfico
        event_data = {
            'id': event.id,
            'theme': event.theme,
            'date': event.date.strftime('%d/%m/%Y'),
            'cities': {
                'labels': list(cities_count.keys()) if cities_count else ['Sem participantes'],
                'data': list(cities_count.values()) if cities_count else [0]
            }
        }
        events_data.append(event_data)

    return render_template(
        'admin_dashboard.html',
        stats=stats,
        recent_events=recent_events,
        events_data=events_data
    )

@app.route('/admin/events', methods=['GET', 'POST'])
@login_required
def manage_events():
    form = EventForm()
    if form.validate_on_submit():
        event = Event(
            theme=form.theme.data,
            date=form.date.data,
            description=form.description.data,
            max_participants=form.max_participants.data
        )
        db.session.add(event)
        db.session.flush()  # Para obter o ID do evento
        
        # Processar os encontros
        for meeting_form in form.meetings.data:
            meeting = Meeting(
                title=meeting_form['title'],
                date=meeting_form['date'],
                time=meeting_form['time'],
                description=meeting_form['description'],
                event_id=event.id
            )
            db.session.add(meeting)

        db.session.commit()
        flash('Evento criado com sucesso!', 'success')
        return redirect(url_for('manage_events'))

    events = Event.query.order_by(Event.date).all()
    today = date.today()
    return render_template('manage_events.html', events=events, form=form, today=today)

@app.route('/admin/events/new', methods=['GET', 'POST'])
@login_required
def create_event():
    form = EventForm()
    if form.validate_on_submit():
        event = Event(
            theme=form.theme.data,
            date=form.date.data,
            max_participants=form.max_participants.data,
            description=form.description.data
        )
        db.session.add(event)
        db.session.commit()
        for meeting_form in form.meetings:
            meeting = Meeting(
                title=meeting_form.title.data,
                date=meeting_form.date.data,
                time=meeting_form.time.data,
                description=meeting_form.description.data,
                event_id=event.id
            )
            db.session.add(meeting)
        db.session.commit()
        flash('Evento e encontros criados com sucesso!', 'success')
        return redirect(url_for('manage_events'))
    return render_template('create_event.html', form=form)

@app.route('/admin/events/<int:event_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_event(event_id):
    event = Event.query.get_or_404(event_id)
    form = EventForm(obj=event)
    if form.validate_on_submit():
        event.theme = form.theme.data
        event.date = form.date.data
        event.max_participants = form.max_participants.data
        event.description = form.description.data
        db.session.commit()
        for meeting_form in form.meetings:
            meeting = Meeting(
                title=meeting_form.title.data,
                date=meeting_form.date.data,
                time=meeting_form.time.data,
                description=meeting_form.description.data,
                event_id=event.id
            )
            db.session.add(meeting)
        db.session.commit()
        flash('Evento e encontros atualizados com sucesso!', 'success')
        return redirect(url_for('manage_events'))
    return render_template('edit_event.html', form=form)

@app.route('/admin/events/<int:event_id>')
@login_required
def get_event(event_id):
    event = Event.query.get_or_404(event_id)
    meetings = Meeting.query.filter_by(event_id=event_id).all()
    
    return jsonify({
        'id': event.id,
        'theme': event.theme,
        'date': event.date.strftime('%Y-%m-%d'),
        'description': event.description,
        'max_participants': event.max_participants,
        'meetings': [{
            'id': meeting.id,
            'title': meeting.title,
            'date': meeting.date.strftime('%Y-%m-%d'),
            'time': meeting.time.strftime('%H:%M'),
            'description': meeting.description
        } for meeting in meetings]
    })

@app.route('/admin/events/save', methods=['POST'])
@login_required
def save_event():
    try:
        event_id = request.args.get('event_id', type=int)
        if event_id:
            event = Event.query.get_or_404(event_id)
        else:
            event = Event()

        # Atualizar dados básicos do evento
        event.theme = request.form['theme']
        event.description = request.form['description']
        event.date = datetime.strptime(request.form['date'], '%Y-%m-%d')
        event.max_participants = request.form.get('max_participants', type=int)

        if not event_id:
            db.session.add(event)
            db.session.flush()  # Para obter o ID do evento antes de adicionar os encontros
        
        # Processar encontros
        meeting_data = []
        i = 0
        while True:
            title = request.form.get(f'meetings-{i}-title')
            if title is None:
                break
                
            date = request.form.get(f'meetings-{i}-date')
            time = request.form.get(f'meetings-{i}-time')
            description = request.form.get(f'meetings-{i}-description')
            
            if title and date and time:
                meeting_data.append({
                    'title': title,
                    'date': datetime.strptime(date, '%Y-%m-%d').date(),
                    'time': datetime.strptime(time, '%H:%M').time(),
                    'description': description
                })
            i += 1

        # Atualizar ou criar encontros
        if event_id:
            # Para edição, remover encontros antigos
            Meeting.query.filter_by(event_id=event.id).delete()
        
        # Adicionar novos encontros
        for data in meeting_data:
            meeting = Meeting(
                event_id=event.id,
                title=data['title'],
                date=data['date'],
                time=data['time'],
                description=data['description']
            )
            db.session.add(meeting)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Evento salvo com sucesso!',
            'event': {
                'id': event.id,
                'theme': event.theme,
                'date': event.date.strftime('%Y-%m-%d'),
                'description': event.description,
                'max_participants': event.max_participants
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Erro ao salvar evento: {str(e)}'
        }), 400

@app.route('/admin/events/delete/<int:event_id>', methods=['POST'])
@login_required
def delete_event(event_id):
    event = Event.query.get_or_404(event_id)
    db.session.delete(event)
    db.session.commit()
    flash('Evento excluído com sucesso!', 'success')
    return redirect(url_for('manage_events'))

@app.route('/admin/manage_participants')
@login_required
def manage_participants():
    participants = Participant.query.all()
    events = Event.query.all()
    form = FlaskForm()  # Para o token CSRF
    return render_template('manage_participants.html', participants=participants, events=events, form=form)

@app.route('/add-participant', methods=['POST'])
@login_required
def add_participant():
    try:
        print("Iniciando add_participant")
        print(f"Método da requisição: {request.method}")
        print(f"Content-Type: {request.headers.get('Content-Type')}")
        print(f"Formulário: {request.form}")
        
        # Obter dados do formulário
        name = request.form.get('full_name')  # Obtém full_name do formulário
        email = request.form.get('email')
        phone = request.form.get('phone')
        city = request.form.get('city')
        event_ids = request.form.getlist('events')
        
        print(f"Dados extraídos:")
        print(f"name: {name}")
        print(f"email: {email}")
        print(f"phone: {phone}")
        print(f"city: {city}")
        print(f"event_ids: {event_ids}")
        
        # Validar dados obrigatórios
        if not all([name, email, phone, city, event_ids]):
            missing = []
            if not name: missing.append('full_name')
            if not email: missing.append('email')
            if not phone: missing.append('phone')
            if not city: missing.append('city')
            if not event_ids: missing.append('events')
            return jsonify({
                'success': False, 
                'message': f'Campos obrigatórios faltando: {", ".join(missing)}'
            })
        
        # Verificar se há pelo menos um evento selecionado
        if not event_ids:
            return jsonify({'success': False, 'message': 'Selecione pelo menos um evento'})
        
        # Verificar se o email já existe
        if Participant.query.filter_by(email=email).first():
            return jsonify({'success': False, 'message': 'Este email já está cadastrado'})
        
        # Verificar se todos os eventos existem
        events = []
        for event_id in event_ids:
            print(f"Procurando evento {event_id}")
            event = Event.query.get(event_id)
            if not event:
                return jsonify({'success': False, 'message': f'Evento {event_id} não encontrado'})
            events.append(event)
        
        # Criar novo participante
        participant = Participant(
            name=name,  # Usa o campo correto do modelo
            email=email,
            phone=phone,
            city=city
        )
        
        # Associar aos eventos selecionados
        for event in events:
            participant.events.append(event)
        
        print("Salvando no banco de dados...")
        # Adicionar ao banco de dados
        db.session.add(participant)
        db.session.commit()
        print("Salvo com sucesso!")
        
        return jsonify({
            'success': True,
            'message': f'Participante cadastrado com sucesso em {len(events)} evento(s)'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Erro ao cadastrar participante: {str(e)}")
        print(f"Tipo do erro: {type(e)}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/participants/<int:participant_id>', methods=['GET'])
@login_required
def get_participant(participant_id):
    participant = Participant.query.get_or_404(participant_id)
    return jsonify({
        'id': participant.id,
        'full_name': participant.full_name,
        'email': participant.email,
        'phone': participant.phone,
        'city': participant.city,
        'events': [event.id for event in participant.events]
    })

@app.route('/admin/participants/<int:participant_id>', methods=['PUT'])
@login_required
def update_participant(participant_id):
    print(f"Iniciando atualização do participante ID: {participant_id}")
    try:
        data = request.get_json()
        print(f"Dados recebidos para atualização: {data}")
        
        # Validar dados obrigatórios
        required_fields = ['full_name', 'email', 'phone', 'city', 'event_ids']
        if not all(field in data for field in required_fields):
            missing_fields = [field for field in required_fields if field not in data]
            error_msg = f"Campos obrigatórios faltando: {', '.join(missing_fields)}"
            print(f"Erro: {error_msg}")
            return jsonify({'success': False, 'message': error_msg})
        
        # Verificar se há pelo menos um evento selecionado
        if not data['event_ids']:
            print("Erro: Nenhum evento selecionado")
            return jsonify({'success': False, 'message': 'Selecione pelo menos um evento'})
        
        # Buscar o participante
        participant = Participant.query.get(participant_id)
        if not participant:
            print(f"Erro: Participante {participant_id} não encontrado")
            return jsonify({'success': False, 'message': 'Participante não encontrado'})
        
        # Verificar se o email ou celular já existem (excluindo o participante atual)
        existing_participant = Participant.query.filter(
            (Participant.email == data['email']) | (Participant.phone == data['phone']),
            Participant.id != participant_id
        ).first()
        
        if existing_participant:
            conflict_type = []
            if existing_participant.email == data['email']:
                conflict_type.append('email')
            if existing_participant.phone == data['phone']:
                conflict_type.append('telefone')
            error_msg = f"Este {' e '.join(conflict_type)} já está cadastrado para outro participante"
            print(f"Erro: {error_msg}")
            return jsonify({'success': False, 'message': error_msg})
        
        # Verificar se todos os eventos existem
        events = []
        for event_id in data['event_ids']:
            event = Event.query.get(event_id)
            if not event:
                error_msg = f'Evento {event_id} não encontrado'
                print(f"Erro: {error_msg}")
                return jsonify({'success': False, 'message': error_msg})
            events.append(event)
        
        # Atualizar dados do participante
        print("Atualizando dados do participante...")
        participant.full_name = data['full_name']
        participant.email = data['email']
        participant.phone = data['phone']
        participant.city = data['city']
        
        # Atualizar eventos
        print(f"Atualizando eventos do participante. Eventos selecionados: {[e.id for e in events]}")
        participant.events = events
        
        # Salvar no banco de dados
        try:
            print("Tentando salvar alterações no banco de dados...")
            db.session.commit()
            print("Alterações salvas com sucesso!")
            return jsonify({
                'success': True,
                'message': f'Participante atualizado com sucesso e associado a {len(events)} evento(s)'
            })
        except Exception as db_error:
            db.session.rollback()
            error_msg = f"Erro ao salvar no banco de dados: {str(db_error)}"
            print(f"Erro: {error_msg}")
            return jsonify({'success': False, 'message': error_msg})
        
    except Exception as e:
        print(f"Erro inesperado ao atualizar participante: {str(e)}")
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro ao atualizar participante: {str(e)}'})

@app.route('/admin/participants/save', methods=['POST'])
@login_required
def save_participant():
    participant_id = request.args.get('participant_id', type=int)
    if participant_id:
        participant = Participant.query.get_or_404(participant_id)
    else:
        participant = Participant()

    participant.full_name = request.form['full_name']
    participant.email = request.form['email']
    participant.phone = request.form['phone']
    participant.city = request.form['city']

    # Atualizar eventos do participante
    event_ids = request.form.getlist('events')
    participant.events = Event.query.filter(Event.id.in_(event_ids)).all()

    if not participant_id:
        db.session.add(participant)
    
    db.session.commit()
    flash('Participante salvo com sucesso!', 'success')
    return redirect(url_for('manage_participants'))

@app.route('/admin/participants/delete/<int:participant_id>', methods=['POST'])
@login_required
def delete_participant(participant_id):
    form = FlaskForm()
    if form.validate_on_submit():
        participant = Participant.query.get_or_404(participant_id)
        db.session.delete(participant)
        db.session.commit()
        flash('Participante excluído com sucesso!', 'success')
    else:
        flash('Erro ao excluir participante. Por favor, tente novamente.', 'error')
    return redirect(url_for('manage_participants'))

@app.route('/admin/meetings')
@login_required
def manage_meetings():
    """Gerenciar encontros."""
    form = MeetingForm()
    events = Event.query.all()
    form.event_id.choices = [(event.id, event.theme) for event in events]

    # Agrupar encontros por evento
    meetings_by_event = {}
    for event in events:
        meetings = Meeting.query.filter_by(event_id=event.id).all()
        for meeting in meetings:
            meeting.registered_count = db.session.query(Attendance).filter_by(meeting_id=meeting.id).count()
        meetings_by_event[event] = meetings

    return render_template('manage_meetings.html', 
                         form=form,
                         meetings_by_event=meetings_by_event)

@app.route('/event/<int:event_id>/meetings', methods=['GET', 'POST'])
@login_required
def manage_event_meetings(event_id):
    event = Event.query.get_or_404(event_id)
    form = MeetingForm()
    if form.validate_on_submit():
        print(f"Título: {form.title.data}, Data: {form.date.data}, Hora: {form.time.data}, Descrição: {form.description.data}")  # Log para depuração
        meeting = Meeting(
            title=form.title.data,
            date=form.date.data,
            time=form.time.data,
            description=form.description.data,
            event_id=event.id
        )
        db.session.add(meeting)
        db.session.commit()
        flash('Encontro adicionado com sucesso!', 'success')
        return redirect(url_for('manage_event_meetings', event_id=event.id))
    meetings = Meeting.query.filter_by(event_id=event_id).all()
    return render_template('manage_meetings.html', event=event, meetings=meetings, form=form)

@app.route('/event/<int:event_id>/meeting/<int:meeting_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_meeting(event_id, meeting_id):
    meeting = Meeting.query.get_or_404(meeting_id)
    form = MeetingForm(obj=meeting)
    if form.validate_on_submit():
        meeting.title = form.title.data
        meeting.date = form.date.data
        meeting.time = form.time.data
        meeting.description = form.description.data
        db.session.commit()
        flash('Encontro atualizado com sucesso!', 'success')
        return redirect(url_for('manage_event_meetings', event_id=event_id))
    return render_template('edit_meeting.html', form=form, meeting=meeting)

@app.route('/admin/meetings/delete/<int:meeting_id>', methods=['POST'])
@login_required
def delete_meeting_admin(meeting_id):
    try:
        meeting = Meeting.query.get_or_404(meeting_id)
        db.session.delete(meeting)
        db.session.commit()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': 'Encontro excluído com sucesso!'
            })
        return redirect(url_for('manage_meetings'))
    except Exception as e:
        db.session.rollback()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': False,
                'message': f'Erro ao excluir encontro: {str(e)}'
            }), 400
        flash('Erro ao excluir encontro', 'error')
        return redirect(url_for('manage_meetings'))

@app.route('/event/<int:event_id>/meeting/<int:meeting_id>/delete', methods=['POST'])
@login_required
def delete_meeting_event(event_id, meeting_id):
    try:
        meeting = Meeting.query.get_or_404(meeting_id)
        db.session.delete(meeting)
        db.session.commit()
        flash('Encontro excluído com sucesso!', 'success')
        return redirect(url_for('manage_event_meetings', event_id=event_id))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir encontro: {str(e)}', 'error')
        return redirect(url_for('manage_event_meetings', event_id=event_id))

@app.route('/admin/meetings/<int:meeting_id>')
@login_required
def get_meeting(meeting_id):
    try:
        meeting = Meeting.query.get_or_404(meeting_id)
        return jsonify({
            'id': meeting.id,
            'event_id': meeting.event_id,
            'title': meeting.title,
            'date': meeting.date.strftime('%Y-%m-%d'),
            'time': meeting.time.strftime('%H:%M'),
            'description': meeting.description or ''
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao carregar encontro: {str(e)}'
        }), 400

@app.route('/admin/meetings/save', methods=['POST'])
@login_required
def save_meeting():
    form = MeetingForm(request.form)
    events = Event.query.all()
    form.event_id.choices = [(event.id, event.theme) for event in events]
    
    if form.validate():
        try:
            meeting_id = request.args.get('meeting_id', type=int)
            if meeting_id:
                meeting = Meeting.query.get_or_404(meeting_id)
            else:
                meeting = Meeting()
            
            # Atualizar dados do encontro
            meeting.event_id = form.event_id.data
            meeting.title = form.title.data
            meeting.date = form.date.data
            meeting.time = form.time.data
            meeting.description = form.description.data
            
            if not meeting_id:
                db.session.add(meeting)
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Encontro salvo com sucesso!',
                'meeting': {
                    'id': meeting.id,
                    'event_id': meeting.event_id,
                    'title': meeting.title,
                    'date': meeting.date.strftime('%Y-%m-%d'),
                    'time': meeting.time.strftime('%H:%M'),
                    'description': meeting.description
                }
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({
                'success': False,
                'message': f'Erro ao salvar encontro: {str(e)}'
            }), 400
    
    return jsonify({
        'success': False,
        'message': 'Erro de validação no formulário',
        'errors': form.errors
    }), 400

@app.route('/admin/meetings/delete/<int:meeting_id>', methods=['POST'])
@login_required
def delete_meeting(meeting_id):
    try:
        meeting = Meeting.query.get_or_404(meeting_id)
        db.session.delete(meeting)
        db.session.commit()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': 'Encontro excluído com sucesso!'
            })
        return redirect(url_for('manage_meetings'))
    except Exception as e:
        db.session.rollback()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': False,
                'message': f'Erro ao excluir encontro: {str(e)}'
            }), 400
        flash('Erro ao excluir encontro', 'error')
        return redirect(url_for('manage_meetings'))

@app.route('/checkin', methods=['GET', 'POST'])
def public_checkin():
    form = CheckinForm()
    now = datetime.now()
    upcoming_event = Event.query.filter(Event.date >= now.date()).order_by(Event.date).first()
    if not upcoming_event:
        flash('Nenhum evento disponível para check-in.', 'error')
        return redirect(url_for('home'))

    event_start = datetime.combine(upcoming_event.date, datetime.min.time())
    checkin_open = event_start - timedelta(hours=3)
    checkin_close = event_start + timedelta(hours=1, minutes=30)

    if now < checkin_open or now > checkin_close:
        flash('Check-in não está disponível no momento.', 'error')
        return redirect(url_for('home'))

    if form.validate_on_submit():
        participant = Participant.query.filter_by(email=form.email.data).first()
        if participant:
            participant.check_in_status = True
            db.session.commit()
            flash('Check-in realizado com sucesso! Bem-vindo ao evento.', 'success')
        else:
            flash('Participante não encontrado. Verifique se o email está correto.', 'error')
        return redirect(url_for('home'))

    return render_template('public_checkin.html', form=form, event=upcoming_event)

@app.route('/admin/export_participants')
@login_required
def export_participants():
    """Exporta a lista de participantes para um arquivo XLSX."""
    try:
        # Criar um novo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Participantes"
        
        # Adicionar cabeçalhos
        headers = ['Nome', 'Email', 'Telefone', 'Cidade', 'Eventos']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Adicionar dados dos participantes
        row = 2
        for participant in Participant.query.all():
            events = ', '.join([event.theme for event in participant.events])
            ws.append([
                participant.full_name,
                participant.email,
                participant.phone,
                participant.city,
                events
            ])
            row += 1
        
        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Salvar em um buffer de memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='participantes.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/download_template')
@login_required
def download_template():
    """Gera um arquivo XLSX modelo para importação de participantes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Cabeçalhos
    headers = ['Nome Completo', 'Email', 'Telefone', 'Cidade', 'ID do Evento']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    # Lista de eventos disponíveis
    ws.append([])  # Linha em branco
    ws.append(['Eventos Disponíveis:'])
    ws.append(['ID', 'Nome do Evento'])
    
    events = Event.query.all()
    for event in events:
        ws.append([event.id, event.theme])
    
    # Configurar largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Salvar em um buffer de memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='template_participantes.xlsx'
    )

@app.route('/admin/import_participants', methods=['POST'])
@login_required
def import_participants():
    """Importa participantes de um arquivo XLSX."""
    if 'file' not in request.files:
        flash('Nenhum arquivo enviado', 'error')
        return redirect(url_for('manage_participants'))
    
    file = request.files['file']
    if file.filename == '':
        flash('Nenhum arquivo selecionado', 'error')
        return redirect(url_for('manage_participants'))
    
    if not file.filename.endswith('.xlsx'):
        flash('Formato de arquivo inválido. Use um arquivo XLSX.', 'error')
        return redirect(url_for('manage_participants'))
    
    try:
        # Ler o arquivo XLSX
        df = pd.read_excel(file)
        required_columns = ['Nome Completo', 'Email', 'Telefone', 'Cidade', 'ID do Evento']
        
        # Verificar se todas as colunas necessárias estão presentes
        if not all(col in df.columns for col in required_columns):
            flash('Arquivo inválido. Certifique-se de usar o template fornecido.', 'error')
            return redirect(url_for('manage_participants'))
        
        success_count = 0
        error_count = 0
        
        for _, row in df.iterrows():
            try:
                # Verificar se o email já existe
                if Participant.query.filter_by(email=row['Email']).first():
                    error_count += 1
                    continue
                
                # Verificar se o evento existe
                event = Event.query.get(row['ID do Evento'])
                if not event:
                    error_count += 1
                    continue
                
                # Criar novo participante
                participant = Participant(
                    full_name=row['Nome Completo'],
                    email=row['Email'],
                    phone=row['Telefone'],
                    city=row['Cidade']
                )
                
                # Adicionar participante ao evento
                participant.events.append(event)
                
                # Adicionar participante a todos os encontros do evento
                for meeting in event.meetings:
                    participant.meetings.append(meeting)
                
                db.session.add(participant)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                continue
        
        db.session.commit()
        
        if success_count > 0:
            flash(f'{success_count} participante(s) importado(s) com sucesso!', 'success')
        if error_count > 0:
            flash(f'{error_count} participante(s) não puderam ser importados.', 'warning')
            
    except Exception as e:
        flash('Erro ao processar o arquivo. Certifique-se de usar o template correto.', 'error')
        
    return redirect(url_for('manage_participants'))

@app.route('/admin/meetings/<int:meeting_id>/attendance')
@login_required
def meeting_attendance_report(meeting_id):
    meeting = Meeting.query.get_or_404(meeting_id)
    
    # Obter todos os participantes do evento
    event_participants = meeting.event.participants
    
    # Criar lista de participantes com status de presença
    attendance_data = []
    for participant in event_participants:
        attendance_data.append({
            'participant': participant,
            'attended': participant in meeting.attendees,
            'attended_at': next((a.attended_at for a in meeting.attendees if a.id == participant.id), None)
        })
    
    # Ordenar por nome
    attendance_data.sort(key=lambda x: x['participant'].full_name)
    
    return render_template('meeting_attendance_report.html',
                         meeting=meeting,
                         attendance_data=attendance_data)

@app.route('/admin/meetings/<int:meeting_id>/attendance/<int:participant_id>', methods=['POST'])
@login_required
def toggle_attendance(meeting_id, participant_id):
    meeting = Meeting.query.get_or_404(meeting_id)
    participant = Participant.query.get_or_404(participant_id)
    
    # Verificar se o participante está inscrito no evento
    if participant not in meeting.event.participants:
        return jsonify({
            'success': False,
            'message': 'Participante não está inscrito neste evento'
        })
    
    try:
        if participant in meeting.attendees:
            # Remover presença
            meeting.attendees.remove(participant)
            message = 'Presença removida com sucesso'
        else:
            # Marcar presença
            meeting.attendees.append(participant)
            message = 'Presença registrada com sucesso'
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': message,
            'attended': participant in meeting.attendees,
            'registered_count': meeting.registered_count,
            'attendance_count': meeting.attendance_count,
            'attendance_percentage': meeting.attendance_percentage
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Erro ao processar presença: {str(e)}'
        })

@app.context_processor
def inject_menu():
    return dict(admin_menu=get_admin_menu())

@app.context_processor
def inject_menu_context():
    def get_menu_items():
        menu_items = []
        
        # Itens básicos do menu que aparecem sempre
        menu_items.append({
            'url': url_for('home'),
            'text': 'Home',
            'icon': 'bi-house'
        })
        
        # Se estiver em uma página de evento específico
        event_id = request.view_args.get('event_id')
        if event_id:
            event = Event.query.get(event_id)
            if event:
                menu_items.extend([
                    {
                        'url': url_for('view_event', event_id=event_id),
                        'text': f'Evento: {event.theme}',
                        'icon': 'bi-calendar-event'
                    },
                    {
                        'url': url_for('manage_event_meetings', event_id=event_id),
                        'text': 'Encontros',
                        'icon': 'bi-clock'
                    },
                    {
                        'url': url_for('manage_participants', event_id=event_id),
                        'text': 'Participantes',
                        'icon': 'bi-people'
                    }
                ])
        else:
            # Menu para área administrativa geral
            menu_items.extend([
                {
                    'url': url_for('manage_events'),
                    'text': 'Eventos',
                    'icon': 'bi-calendar3'
                },
                {
                    'url': url_for('manage_meetings'),
                    'text': 'Encontros',
                    'icon': 'bi-clock'
                }
            ])
        
        return menu_items
    
    return dict(get_menu_items=get_menu_items)

@app.context_processor
def inject_user():
    """Injeta o usuário atual em todos os templates"""
    return {
        'current_user': {
            'is_authenticated': 'logged_in' in session,
            'username': session.get('username', 'Admin')
        }
    }

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True)
